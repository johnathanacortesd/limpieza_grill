# ======================================
# Importaciones
# ======================================
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from collections import defaultdict, Counter
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import re
import time
from unidecode import unidecode
import numpy as np
import gc
from pathlib import Path

# ======================================
# Configuración general
# ======================================
st.set_page_config(
    page_title="Limpieza de Noticias · Realizado por Johnathan Cortés",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

SIMILARITY_THRESHOLD_TITULOS = 0.93

# ======================================
# CSS Personalizado
# ======================================
def load_custom_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Google+Sans:wght@400;500;700&family=Google+Sans+Text:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
:root {
    --bg:#f8f9fa;--s1:#ffffff;--s2:#f1f3f4;--s3:#e8eaed;
    --border:#dadce0;--border2:#bdc1c6;--border-focus:#f97316;
    --text:#202124;--text2:#3c4043;--text3:#5f6368;--text4:#9aa0a6;
    --accent:#f97316;--accent2:#ea580c;--accent3:#c2410c;
    --accent-bg:#fff7ed;--accent-bg2:#ffedd5;--accent-bdr:#fed7aa;
    --green:#059669;--green2:#047857;--green-bg:#ecfdf5;--green-bdr:#a7f3d0;
    --red:#dc2626;--amber:#d97706;--blue:#1a73e8;
    --r:8px;--r2:12px;--r3:16px;--r4:20px;
    --shadow-sm:0 1px 2px rgba(60,64,67,0.1),0 1px 3px rgba(60,64,67,0.08);
    --shadow-md:0 1px 3px rgba(60,64,67,0.12),0 4px 8px rgba(60,64,67,0.08);
    --shadow-lg:0 2px 6px rgba(60,64,67,0.1),0 8px 24px rgba(60,64,67,0.1);
    --transition:all 0.2s cubic-bezier(0.4,0,0.2,1);
}
html,body,[data-testid="stApp"]{
    background:var(--bg)!important;color:var(--text)!important;
    font-family:'Google Sans Text','Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
    font-size:14px;-webkit-font-smoothing:antialiased;letter-spacing:0.01em;
}
#MainMenu,footer,header{visibility:hidden}.stDeployButton{display:none}
.block-container{padding-top:1rem!important;padding-bottom:0!important}
[data-testid="stAppViewBlockContainer"]{padding-top:1rem!important}
.app-header{background:var(--s1);border:1px solid var(--border);border-radius:var(--r3);padding:1rem 1.5rem;margin-bottom:1rem;display:flex;align-items:center;gap:1rem;box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.app-header::after{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,#f97316,#fb923c,#fdba74);}
.app-header-icon{width:40px;height:40px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;color:white;flex-shrink:0;box-shadow:0 2px 8px rgba(249,115,22,0.3);}
.app-header-text{flex:1}
.app-header-title{font-family:'Google Sans',sans-serif;font-size:1.25rem;font-weight:700;color:var(--text);letter-spacing:-0.01em;line-height:1.3}
.app-header-version{font-family:'Roboto Mono',monospace;font-size:0.65rem;color:var(--text3);letter-spacing:0.03em;margin-top:0.15rem}
.app-header-badge{background:var(--accent-bg);border:1px solid var(--accent-bdr);color:var(--accent2);font-family:'Roboto Mono',monospace;font-size:0.6rem;font-weight:500;padding:0.25rem 0.75rem;border-radius:100px;letter-spacing:0.04em;text-transform:uppercase;white-space:nowrap;}
.metrics-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:0.6rem;margin:0.8rem 0}
.metric-card{background:var(--s1);border:1px solid var(--border);border-radius:var(--r2);padding:0.8rem 0.6rem;text-align:center;transition:var(--transition);box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.metric-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--r2) var(--r2) 0 0}
.metric-card.m-total::before{background:linear-gradient(90deg,#5f6368,#9aa0a6)}
.metric-card.m-unique::before{background:linear-gradient(90deg,#059669,#34d399)}
.metric-card.m-dup::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.metric-card.m-time::before{background:linear-gradient(90deg,#1a73e8,#4285f4)}
.metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-lg)}
.metric-val{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;line-height:1;margin-bottom:0.3rem;letter-spacing:-0.01em}
.metric-lbl{font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);text-transform:uppercase;letter-spacing:0.08em;font-weight:500}
[data-testid="stForm"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r3)!important;padding:1.2rem 1.5rem!important;box-shadow:var(--shadow-md)!important;}
.sec-label{font-family:'Google Sans',sans-serif;font-size:0.72rem;font-weight:700;color:var(--text2);letter-spacing:0.08em;text-transform:uppercase;padding-bottom:0.3rem;border-bottom:2px solid var(--s3);margin:0.8rem 0 0.5rem;display:flex;align-items:center;gap:0.5rem;}
.sec-label::before{content:'';display:inline-block;width:3px;height:12px;background:linear-gradient(180deg,#f97316,#ea580c);border-radius:2px}
.upload-zone{display:grid;grid-template-columns:1fr;gap:0.6rem;margin:0.3rem 0}
.upload-zone-card{background:var(--s1);border:1.5px dashed var(--border);border-radius:var(--r2);padding:0.6rem 0.8rem;display:flex;align-items:center;gap:0.6rem;transition:var(--transition);}
.upload-zone-card:hover{border-color:var(--accent);border-style:solid;transform:translateY(-1px);box-shadow:var(--shadow-md)}
.upload-zone-icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1rem;flex-shrink:0;}
.upload-zone-icon.uz-dossier{background:#fff7ed;color:#f97316}
.upload-zone-text{flex:1;min-width:0}
.upload-zone-title{font-family:'Google Sans',sans-serif;font-size:0.82rem;font-weight:700;color:var(--text);line-height:1.2}
.upload-zone-desc{font-size:0.7rem;color:var(--text3);line-height:1.3}
[data-testid="stFileUploader"]{background:var(--s1)!important;border:1.5px dashed var(--border)!important;border-radius:var(--r)!important;padding:0.4rem 0.6rem!important;transition:var(--transition)!important;min-height:auto!important;}
[data-testid="stFileUploader"]:hover{border-color:var(--accent)!important;border-style:solid!important;background:var(--accent-bg)!important;}
[data-testid="stFileUploader"] section{padding:0.2rem!important}
[data-testid="stFileUploader"] section>div{font-size:0.78rem!important;color:var(--text2)!important}
[data-testid="stFileUploader"] section small{font-size:0.7rem!important;color:var(--text3)!important}
[data-testid="stFileUploader"] button{background:var(--accent-bg)!important;border:1px solid var(--accent-bdr)!important;color:var(--accent2)!important;font-weight:500!important;font-size:0.75rem!important;border-radius:100px!important;padding:0.25rem 0.8rem!important;font-family:'Google Sans',sans-serif!important;transition:var(--transition)!important;}
[data-testid="stFileUploader"] button:hover{background:var(--accent)!important;color:white!important;border-color:var(--accent)!important}
[data-testid="stTextInput"] input{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:var(--r)!important;font-family:'Google Sans Text',sans-serif!important;font-size:0.9rem!important;padding:0.5rem 0.75rem!important;transition:var(--transition)!important;}
[data-testid="stTextInput"] input:focus{border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(249,115,22,0.12)!important;}
label[data-testid="stWidgetLabel"] p{font-family:'Google Sans',sans-serif!important;color:var(--text2)!important;font-size:0.82rem!important;font-weight:500!important;margin-bottom:0.15rem!important;}
.stButton>button,[data-testid="stDownloadButton"]>button{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:100px!important;font-family:'Google Sans',sans-serif!important;font-weight:500!important;font-size:0.88rem!important;transition:var(--transition)!important;padding:0.5rem 1.2rem!important;box-shadow:none!important;}
.stButton>button:hover,[data-testid="stDownloadButton"]>button:hover{border-color:var(--accent)!important;color:var(--accent2)!important;background:var(--accent-bg)!important;box-shadow:var(--shadow-sm)!important;transform:translateY(-1px)!important;}
.stButton>button[kind="primary"],[data-testid="stDownloadButton"]>button[kind="primary"]{background:var(--accent)!important;border:none!important;color:#fff!important;font-weight:500!important;font-size:0.92rem!important;padding:0.6rem 1.5rem!important;box-shadow:0 1px 3px rgba(249,115,22,0.3),0 4px 12px rgba(249,115,22,0.15)!important;letter-spacing:0.01em!important;}
.stButton>button[kind="primary"]:hover,[data-testid="stDownloadButton"]>button[kind="primary"]:hover{background:var(--accent2)!important;box-shadow:0 2px 6px rgba(234,88,12,0.35),0 8px 24px rgba(234,88,12,0.18)!important;transform:translateY(-1px)!important;color:#fff!important;}
.success-banner{background:linear-gradient(135deg,#ecfdf5,#d1fae5);border:1px solid var(--green-bdr);border-left:4px solid var(--green);border-radius:var(--r2);padding:0.8rem 1.2rem;margin:0.5rem 0 0.8rem;display:flex;align-items:center;gap:0.8rem;}
.success-icon{width:34px;height:34px;background:linear-gradient(135deg,#059669,#047857);border-radius:50%;display:flex;align-items:center;justify-content:center;color:white;font-size:1rem;flex-shrink:0;}
.success-title{font-family:'Google Sans',sans-serif;font-size:1rem;font-weight:700;color:#047857;margin-bottom:0.1rem}
.success-sub{font-size:0.8rem;color:var(--text2)}
.auth-wrap{max-width:380px;margin:8vh auto 0;text-align:center}
.auth-icon{width:60px;height:60px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:16px;display:inline-flex;align-items:center;justify-content:center;font-size:1.6rem;color:white;margin-bottom:1rem;box-shadow:0 4px 166px rgba(249,115,22,0.3);}
.auth-title{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;color:var(--text);margin-bottom:0.3rem}
.auth-sub{font-size:0.85rem;color:var(--text3);margin-bottom:2rem}
[data-testid="stProgressBar"]>div>div{background:linear-gradient(90deg,#f97316,#fb923c,#fdba74)!important;border-radius:100px!important;height:5px!important;}
[data-testid="stDataFrame"]{border:1px solid var(--border)!important;border-radius:var(--r2)!important;box-shadow:var(--shadow-sm)!important;overflow:hidden!important;}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:var(--s2);border-radius:3px}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--accent)}
.footer{font-family:'Roboto Mono',monospace;font-size:0.6rem;color:var(--text4);text-align:center;padding:0.8rem 0 0.5rem;letter-spacing:0.04em;border-top:1px solid var(--s3);margin-top:1rem;}
.stElementContainer{margin-bottom:0!important}
[data-testid="stVerticalBlock"]>div{gap:0.3rem!important}
[data-testid="stHorizontalBlock"]>div{gap:0.4rem!important}
hr{border-color:var(--s3)!important;margin:0.5rem 0!important}
@media(max-width:768px){
    .metrics-grid{grid-template-columns:repeat(2,1fr)}
    .app-header{flex-direction:column;text-align:center;gap:0.5rem;padding:1rem}
}
</style>
""", unsafe_allow_html=True)

# ======================================
# Autenticación Básica
# ======================================
def check_password():
    if st.session_state.get("password_correct", False):
        return True
    st.markdown("""
    <div class="auth-wrap">
        <div class="auth-icon">◈</div>
        <div class="auth-title">Sistema de Limpieza</div>
        <div class="auth-sub">Ingresa tus credenciales para continuar</div>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("pw"):
            pw = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            if st.form_submit_button("Ingresar", use_container_width=True, type="primary"):
                if pw == st.secrets.get("APP_PASSWORD", "INVALID"):
                    st.session_state["password_correct"] = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
    return False

# ======================================
# Mapeos de Configuración Local
# ======================================
def load_local_config():
    paths_to_try = [
        Path("Configuracion.xlsx"),
        Path("configuracion.xlsx"),
        Path("Config.xlsx"),
        Path("config.xlsx")
    ]
    for p in paths_to_try:
        if p.exists():
            return p
    base = Path(__file__).parent
    for f in base.iterdir():
        if f.suffix.lower() == '.xlsx' and 'config' in f.stem.lower():
            return f
    return None

def load_config(config_source):
    config_sheets = pd.read_excel(config_source, sheet_name=None, engine='openpyxl')
    region_map = pd.Series(
        config_sheets['Regiones'].iloc[:, 1].values,
        index=config_sheets['Regiones'].iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    internet_map = pd.Series(
        config_sheets['Internet'].iloc[:, 1].values,
        index=config_sheets['Internet'].iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    return region_map, internet_map

# ======================================
# Utilidades de Limpieza de Texto
# ======================================
def norm_key(text):
    if text is None: return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))

def get_column_robust(df, name):
    """Busca una columna de forma insensible a mayúsculas, minúsculas, espacios y tildes."""
    name_norm = norm_key(name)
    for col in df.columns:
        if norm_key(col) == name_norm:
            return df[col]
    return pd.Series([np.nan] * len(df))

def clean_text(text):
    if not isinstance(text, str):
        return text
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def clean_cuerpo(text):
    if not isinstance(text, str) or text.strip() == '':
        return text
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()

def normalize_title_for_comparison(title):
    if not isinstance(title, str): return ""
    tmp = re.split(r"\s*[:|-]\s*", title, 1)
    return re.sub(r"\W+", " ", tmp[0]).lower().strip()

def clean_title_for_output(title):
    return re.sub(r"\s*\|\s*[\w\s]+$", "", str(title)).strip()

def corregir_texto(text):
    if not isinstance(text, str): return text
    text = re.sub(r"(<br>|\[\.\.\.\]|\s+)", " ", text).strip()
    m = re.search(r"[A-ZÁÉÍÓÚÑ]", text)
    if m: text = text[m.start():]
    if text and not text.endswith("..."): text = text.rstrip(".") + "..."
    return text

def normalizar_tipo_medio(tipo_raw):
    if not isinstance(tipo_raw, str): return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    return {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio', 'radio': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión', 'tv': 'Televisión',
        'television': 'Televisión', 'televisión': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }.get(t, str(tipo_raw).strip().title() or "Otro")

def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    s = str(val).strip()
    if not s:
        return None
    s = s.replace(',', '')
    try:
        f_val = float(s)
        if f_val.is_integer():
            return int(f_val)
        return f_val
    except ValueError:
        return None

# ======================================
# Transformación de Tono
# ======================================
def mapped_tono(val):
    if pd.isna(val) or val is None:
        return ""
    s = str(val).strip().lower()
    if 'positiv' in s:
        return 'Positivo'
    elif 'neutr' in s:
        return 'Neutro'
    elif 'negativ' in s:
        return 'Negativo'
    if s in ('', 'nan', 'none'):
        return ""
    return str(val).strip()

# ======================================
# Algoritmo de Duplicados Local
# ======================================
def _normalizar_url(url: str) -> str:
    if not url: return ""
    url = url.strip().lower()
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    url = url.rstrip('/')
    return url

def detectar_duplicados_avanzado(rows, km):
    processed = deepcopy(rows)
    seen_url, seen_bcast = {}, {}
    seen_streaming: Dict[tuple, int] = {}
    tb = defaultdict(list)

    for i, row in enumerate(processed):
        if row.get("is_duplicate"): continue

        tipo    = normalizar_tipo_medio(str(row.get(km["tipodemedio"], "")))
        mencion = norm_key(row.get(km["menciones"], ""))
        medio   = norm_key(row.get(km["medio"], ""))

        streaming_url_raw = row.get(km["link_streaming"])
        if isinstance(streaming_url_raw, dict):
            streaming_url_raw = streaming_url_raw.get("url")
            
        if streaming_url_raw and mencion:
            streaming_url_norm = _normalizar_url(str(streaming_url_raw))
            if streaming_url_norm:
                sk = (streaming_url_norm, mencion)
                if sk in seen_streaming:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_streaming[sk]].get(km["idnoticia"], "")
                    continue
                seen_streaming[sk] = i

        if tipo == "Internet":
            li = row.get(km["link_nota"])
            url = li.get("url") if isinstance(li, dict) else li
            if url and mencion:
                url_norm = _normalizar_url(str(url))
                k = (url_norm, mencion)
                if k in seen_url:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_url[k]].get(km["idnoticia"], "")
                    continue
                seen_url[k] = i
            if medio and mencion:
                tb[(medio, mencion)].append(i)

        elif tipo in ("Radio", "Televisión"):
            hora = str(row.get(km["hora"], "")).strip()
            if mencion and medio and hora:
                k = (mencion, medio, hora)
                if k in seen_bcast:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_bcast[k]].get(km["idnoticia"], "")
                else:
                    seen_bcast[k] = i

    for idxs in tb.values():
        if len(idxs) < 2: continue
        for i in range(len(idxs)):
            for j in range(i + 1, len(idxs)):
                a, b = idxs[i], idxs[j]
                if processed[a].get("is_duplicate") or processed[b].get("is_duplicate"): continue
                ta  = normalize_title_for_comparison(processed[a].get(km["titulo"]))
                tb_ = normalize_title_for_comparison(processed[b].get(km["titulo"]))
                if ta and tb_ and SequenceMatcher(None, ta, tb_).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    if len(ta) < len(tb_):
                        processed[a]["is_duplicate"] = True
                        processed[a][km["idduplicada"]]  = processed[b].get(km["idnoticia"], "")
                    else:
                        processed[b]["is_duplicate"] = True
                        processed[b][km["idduplicada"]]  = processed[a].get(km["idnoticia"], "")

    return processed

# ======================================
# Lectura y Estructuración de Datos
# ======================================
def read_and_normalize_dossier(sheet, region_map, internet_map):
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    rows = []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue
        row_data = {}
        for i, h in enumerate(headers):
            if i < len(row):
                cell = row[i]
                val = cell.value
                url = cell.hyperlink.target if (cell.hyperlink and cell.hyperlink.target) else None
                if url:
                    row_data[h] = {"value": val or "Link", "url": url}
                else:
                    row_data[h] = val
        rows.append(row_data)

    df = pd.DataFrame(rows)

    tipo_medio_map = {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }
    
    if 'Tipo de Medio' in df.columns:
        df['Tipo de Medio'] = (
            df['Tipo de Medio'].astype(str).str.lower().str.strip()
            .map(tipo_medio_map)
            .fillna(df['Tipo de Medio'].astype(str).str.strip())
        )
    else:
        df['Tipo de Medio'] = 'Otro'

    is_av = df['Tipo de Medio'].isin(['Radio', 'Televisión'])
    is_grafica = df['Tipo de Medio'].isin(['Prensa', 'Internet', 'Revistas'])
    is_internet = df['Tipo de Medio'] == 'Internet'

    # Captura del texto totalmente original de la columna Resumen - Aclaracion
    raw_resumen_orig = get_column_robust(df, 'Resumen - Aclaracion')

    # Mapeo Región
    if 'Medio' in df.columns:
        raw_medios_clean = df['Medio'].astype(str).str.lower().str.strip()
        df['Región'] = raw_medios_clean.map(region_map).fillna("N/A")
    else:
        df['Medio'] = 'N/A'
        df['Región'] = 'N/A'

    # Mapeo Medio para Internet
    if 'Medio' in df.columns:
        df.loc[is_internet, 'Medio'] = (
            df.loc[is_internet, 'Medio']
            .astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, 'Medio'])
        )

    df['ID Noticia'] = df.get('NoticiaId', df.get('ID Noticia', pd.Series(dtype=str)))
    df['Fecha'] = pd.to_datetime(df.get('Fecha', pd.Series(dtype=str)), dayfirst=True, errors='coerce').dt.normalize()
    df['Hora'] = df.get('Hora', pd.Series(dtype=str))
    df['Sección - Programa'] = df.get('Sección - Programa', pd.Series(dtype=str)).astype(str).apply(clean_text)
    
    titulo_col = 'Título' if 'Título' in df.columns else 'Titulo'
    df['Título'] = df.get(titulo_col, pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Autor - Conductor'] = df.get('Autor - Conductor', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Nro. Pagina'] = df.get('Nro. Pagina', pd.Series(dtype=str))
    
    dim_col = 'Dimensioncm2' if 'Dimensioncm2' in df.columns else 'Dimensión'
    df['Dimensión'] = df.get(dim_col, pd.Series(dtype=str))
    df['Duración - Nro. Caracteres'] = df.get('Duración - Nro. Caracteres', pd.Series(dtype=str))

    df.loc[is_av, 'Dimensión'] = df.loc[is_av, 'Duración - Nro. Caracteres']
    df.loc[is_av, 'Duración - Nro. Caracteres'] = 0

    # Lógica de CPE y Revalorización robusta sin notación científica
    cpe_input = get_column_robust(df, 'CPE')
    valor_nota_input = get_column_robust(df, 'Valor de Nota')

    # CPE resultante en base a las nuevas reglas:
    # Radio o Televisión (is_av) = columna CPE original
    # Internet, Prensa o Revistas (is_grafica) = columna Valor de Nota original
    df['CPE'] = np.where(is_av, cpe_input, np.where(is_grafica, valor_nota_input, np.nan))

    # revalorización resultante: mantiene la regla de que solo para Internet, Prensa o Revistas se asigne la columna CPE original
    df['revalorización'] = np.where(is_grafica, cpe_input, np.nan)

    df['Tier'] = df.get('Tier', pd.Series(dtype=str))
    df['Audiencia'] = df.get('Audiencia', pd.Series(dtype=str))
    
    # Conservar y transformar Tono
    df['Tono'] = get_column_robust(df, 'Tono').apply(mapped_tono)
    
    # Conservar Tematica como Tema, y Subtema
    df['Tema'] = get_column_robust(df, 'Tematica').fillna('').astype(str).apply(clean_text)
    df['Subtema'] = get_column_robust(df, 'Subtema').fillna('').astype(str).apply(clean_text)

    # Traer la información real de las columnas correspondientes del Excel de origen
    df['Producto'] = get_column_robust(df, 'Producto').fillna('').astype(str).apply(clean_text)
    df['Tipo de información'] = get_column_robust(df, 'Tipo de información').fillna('').astype(str).apply(clean_text)
    df['Nombre vocero'] = get_column_robust(df, 'Nombre vocero').fillna('').astype(str).apply(clean_text)
    df['Mención en Titulo'] = get_column_robust(df, 'Mención en Titulo').fillna('').astype(str).apply(clean_text)
    df['Mención en Foto'] = get_column_robust(df, 'Mención en Foto').fillna('').astype(str).apply(clean_text)
    df['Tipo mencion'] = get_column_robust(df, 'Tipo mencion').fillna('').astype(str).apply(clean_text)
    df['Tipo mencion 2'] = get_column_robust(df, 'Tipo mencion 2').fillna('').astype(str).apply(clean_text)
    df['Aparece Logo'] = get_column_robust(df, 'Aparece Logo').fillna('').astype(str).apply(clean_text)

    # Asignar a 'resumen corto' exactamente el texto original que se leyó de forma robusta al inicio
    df['resumen corto'] = raw_resumen_orig.fillna('').astype(str).str.strip()

    # Procesar limpieza normal para la columna final "Resumen - Aclaracion"
    cuerpo_col = 'CuerpoEs' if 'CuerpoEs' in df.columns else 'Resumen - Aclaracion'
    cuerpo_cleaned = df.get(cuerpo_col, pd.Series([''] * len(df))).astype(str).apply(clean_cuerpo)

    def fmt_grafica(text):
        if not isinstance(text, str) or not text.strip():
            return text
        parrafos = [p.strip() for p in text.split('\n') if p.strip()]
        return '\n\n'.join(parrafos) if len(parrafos) > 1 else text

    df['Resumen - Aclaracion'] = np.where(is_av, cuerpo_cleaned, cuerpo_cleaned.apply(fmt_grafica))

    # Links
    url_nota_av = df.get('URL Nota AV', df.get('Link Nota AV', pd.Series([''] * len(df))))
    url_streaming = df.get('URL (Streaming - Imagen)', pd.Series([''] * len(df)))
    
    link_nota_final = []
    for val_av, val_str, is_av_row in zip(url_nota_av, url_streaming, is_av):
        if is_av_row:
            if isinstance(val_av, dict):
                url_t = val_av.get("url", "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
            else:
                url_t = str(val_av or "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
        else:
            if isinstance(val_str, dict):
                link_nota_final.append(val_str)
            else:
                link_nota_final.append({"value": "Link", "url": val_str if val_str else None})
                
    df['Link Nota'] = link_nota_final

    url_nota_raw = df.get('URL Nota', pd.Series([''] * len(df)))
    link_stream_final = []
    for val_url, is_int in zip(url_nota_raw, is_internet):
        if is_int:
            if isinstance(val_url, dict):
                link_stream_final.append(val_url)
            else:
                link_stream_final.append({"value": "Link", "url": val_url if val_url else None})
        else:
            link_stream_final.append(None)
            
    df['Link (Streaming - Imagen)'] = link_stream_final

    menciones_av = df.get('Menciones - Empresa', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    menciones_grafica = df.get('Empresa rel.', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    df['Menciones - Empresa'] = np.where(is_av, menciones_av, np.where(is_grafica, menciones_grafica, menciones_av))

    return df

# ======================================
# Exportar a Excel Estructurado
# ======================================
def generate_output_excel(rows, km):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    
    ORDER = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio",
        "Sección - Programa", "Región", "Título", "Autor - Conductor",
        "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres",
        "CPE", "Tier", "Audiencia", "Tono", "Tema", "Subtema",
        "Producto", "Tipo de información", "Nombre vocero",
        "Mención en Titulo", "Mención en Foto", "Tipo mencion",
        "Tipo mencion 2", "Aparece Logo", "revalorización", "resumen corto",
        "Link Nota", "Resumen - Aclaracion", "Link (Streaming - Imagen)", "Menciones - Empresa",
        "ID duplicada"
    ]
    
    NUM = {"ID Noticia", "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "CPE", "revalorización", "Tier", "Audiencia"}
    ws.append(ORDER)
    
    font_hyperlink = Font(color="0563C1", underline="single")
    align_left = Alignment(horizontal='left')
    font_header = Font(bold=True)
    
    for i, col_name in enumerate(ORDER, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = font_header
        
    for row in rows:
        tk = km.get("titulo")
        if tk and tk in row: row[tk] = clean_title_for_output(row.get(tk))
        rk = km.get("resumen")
        if rk and rk in row: row[rk] = corregir_texto(row.get(rk))
        
        out, links = [], {}
        for ci, h in enumerate(ORDER, start=1):
            val = row.get(h)
            cv = None
            
            if h == 'Fecha' and pd.notna(val):
                if isinstance(val, pd.Timestamp):
                    cv = val.to_pydatetime()
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    cv = val
                else:
                    cv = str(val) if val is not None else None
            elif h in NUM:
                cv = parse_numeric(val)
            elif isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"): links[ci] = val["url"]
            elif val is not None:
                if isinstance(val, str) and val.startswith("http"):
                    cv = "Link"
                    links[ci] = val
                else:
                    cv = str(val)
            out.append(cv)
        ws.append(out)
        
        current_row = ws.max_row
        for ci, url in links.items():
            cell = ws.cell(row=current_row, column=ci)
            cell.hyperlink = url
            cell.font = font_hyperlink
            cell.alignment = align_left
            
        date_col_idx = ORDER.index("Fecha") + 1
        date_cell = ws.cell(row=current_row, column=date_col_idx)
        if isinstance(date_cell.value, (datetime.datetime, datetime.date)):
            date_cell.number_format = 'DD/MM/YYYY'
            
        cpe_idx = ORDER.index("CPE") + 1
        reval_idx = ORDER.index("revalorización") + 1
        
        cpe_cell = ws.cell(row=current_row, column=cpe_idx)
        reval_cell = ws.cell(row=current_row, column=reval_idx)
        
        if isinstance(cpe_cell.value, (int, float)):
            cpe_cell.number_format = '#,##0'
        if isinstance(reval_cell.value, (int, float)):
            reval_cell.number_format = '#,##0'
            
    for i, col_name in enumerate(ORDER, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        if col_name in ['Título', 'Resumen - Aclaracion', 'resumen corto']:
            ws.column_dimensions[letter].width = 50
        elif col_name in ['Link Nota', 'Link (Streaming - Imagen)']:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 20
            
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ======================================
# Proceso Principal
# ======================================
def run_cleaning_process(df_file):
    t0 = time.time()
    
    with st.status("Cargando Configuración y Dossier", expanded=True) as s:
        config_path = load_local_config()
        if not config_path:
            st.error("❌ No se encontró el archivo 'Configuracion.xlsx' en el repositorio.")
            st.stop()
            
        region_map, internet_map = load_config(config_path)
        wb_in = load_workbook(df_file, data_only=True)
        df_normalized = read_and_normalize_dossier(wb_in.active, region_map, internet_map)
        
        # Expansión por punto y coma (;) en Menciones - Empresa
        rows_expanded = []
        for idx, row_series in df_normalized.iterrows():
            menciones = [m.strip() for m in str(row_series['Menciones - Empresa']).split(';') if m.strip()]
            if not menciones:
                row_dict = row_series.to_dict()
                row_dict['Menciones - Empresa'] = ""
                row_dict['original_index'] = idx
                row_dict['is_duplicate'] = False
                rows_expanded.append(row_dict)
            else:
                for m in menciones:
                    row_dict = row_series.to_dict()
                    row_dict['Menciones - Empresa'] = m
                    row_dict['original_index'] = idx
                    row_dict['is_duplicate'] = False
                    rows_expanded.append(row_dict)

        km = {
            "idnoticia": "ID Noticia",
            "fecha": "Fecha",
            "hora": "Hora",
            "medio": "Medio",
            "tipodemedio": "Tipo de Medio",
            "seccion_programa": "Sección - Programa",
            "region": "Región",
            "titulo": "Título",
            "autor_conductor": "Autor - Conductor",
            "nro_pagina": "Nro. Pagina",
            "dimension": "Dimensión",
            "duracion_caracteres": "Duración - Nro. Caracteres",
            "cpe": "CPE",
            "tier": "Tier",
            "audiencia": "Audiencia",
            "tono": "Tono",
            "tema": "Tema",
            "subtema": "Subtema",
            "producto": "Producto",
            "tipo_de_informacion": "Tipo de información",
            "nombre_vocero": "Nombre vocero",
            "mencion_en_titulo": "Mención en Titulo",
            "mencion_en_foto": "Mención en Foto",
            "tipo_mencion": "Tipo mencion",
            "tipo_mencion_2": "Tipo mencion 2",
            "aparece_logo": "Aparece Logo",
            "revalorizacion": "revalorización",
            "resumen_corto": "resumen corto",
            "link_nota": "Link Nota",
            "resumen": "Resumen - Aclaracion",
            "link_streaming": "Link (Streaming - Imagen)",
            "menciones": "Menciones - Empresa",
            "idduplicada": "ID duplicada"
        }
        
        rows = detectar_duplicados_avanzado(rows_expanded, km)
        for row in rows:
            if row["is_duplicate"]:
                row["Tono"] = "Duplicada"
                row["Tema"] = "-"
                row["Subtema"] = "-"
                
        s.update(label="✓ Archivo estructurado con éxito", state="complete")
        
    gc.collect()
    ta = [r for r in rows if not r.get("is_duplicate")]
    
    st.session_state["output_data"] = generate_output_excel(rows, km)
    st.session_state["output_filename"] = f"Dossier_Limpio_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.session_state["processing_complete"] = True
    st.session_state.update({
        "total_rows": len(rows),
        "unique_rows": len(ta),
        "duplicates": len(rows) - len(ta),
        "process_duration": f"{time.time() - t0:.2f}s"
    })

# ======================================
# Interfaz de Usuario
# ======================================
def main():
    load_custom_css()
    if not check_password(): return

    st.markdown("""
    <div class="app-header">
        <div class="app-header-icon">◈</div>
        <div class="app-header-text">
            <div class="app-header-title">Limpieza de Xlsx Grill</div>
            <div class="app-header-version">v2.4 · Realizado por Johnathan Cortés</div>
        </div>
        <div class="app-header-badge">Estructurador</div>
    </div>""", unsafe_allow_html=True)

    if not st.session_state.get("processing_complete", False):
        with st.form("main_form"):
            st.markdown('<div class="sec-label">Sube el archivo de entrada</div>', unsafe_allow_html=True)
            st.markdown("""
            <div class="upload-zone">
                <div class="upload-zone-card">
                    <div class="upload-zone-icon uz-dossier">📋</div>
                    <div class="upload-zone-text">
                        <div class="upload-zone-title">Dossier de Noticias</div>
                        <div class="upload-zone-desc">Sube el archivo .xlsx para aplicar el flujo de normalización</div>
                    </div>
                </div>
            </div>""", unsafe_allow_html=True)
            
            f1 = st.file_uploader("Dossier", type=["xlsx"], label_visibility="collapsed", key="f1")

            if st.form_submit_button("▶ Iniciar Limpieza", use_container_width=True, type="primary"):
                if not f1:
                    st.error("Por favor, sube un archivo Excel.")
                else:
                    run_cleaning_process(f1)
                    st.rerun()
    else:
        total = st.session_state.total_rows
        uniq  = st.session_state.unique_rows
        dups  = st.session_state.duplicates
        dur   = st.session_state.process_duration
        
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Limpieza completada</div>'
            '<div class="success-sub">El archivo estructurado se encuentra listo para descargar</div></div></div>',
            unsafe_allow_html=True
        )
        
        st.markdown(f"""
        <div class="metrics-grid">
          <div class="metric-card m-total"><div class="metric-val" style="color:var(--text)">{total}</div><div class="metric-lbl">Total Registros</div></div>
          <div class="metric-card m-unique"><div class="metric-val" style="color:var(--green)">{uniq}</div><div class="metric-lbl">Únicos</div></div>
          <div class="metric-card m-dup"><div class="metric-val" style="color:var(--amber)">{dups}</div><div class="metric-lbl">Duplicados</div></div>
          <div class="metric-card m-time"><div class="metric-val" style="color:var(--blue)">{dur}</div><div class="metric-lbl">Tiempo de Ejecución</div></div>
        </div>""", unsafe_allow_html=True)
        
        c1, c2 = st.columns(2)
        c1.download_button(
            "⬇ Descargar Xlsx Limpio",
            data=st.session_state.output_data,
            file_name=st.session_state.output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        if c2.button("Nuevo análisis", use_container_width=True):
            pwd = st.session_state.get("password_correct")
            st.session_state.clear()
            st.session_state.password_correct = pwd
            st.rerun()

    st.markdown(
        '<div class="footer">Estructuración y Limpieza · Johnathan Cortés ©</div>',
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
